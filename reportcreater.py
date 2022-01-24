from asyncio.proactor_events import constants
import pandas as pd
import datetime as dt
import sys
import threading
import subprocess

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

import myconstants
from myutils import get_files_list, save_param, load_param

def thread(my_func):
    """
    Запускает функцию в отдельном потоке
    """
    def wrapper(*args, **kwargs):
        my_thread = threading.Thread(target=my_func, args=args, kwargs=kwargs)
        my_thread.start()
    return wrapper

class ReportCreater(object):
    def __init__(self, *args):
        super(ReportCreater, self).__init__(*args)
        self.reports_list = get_files_list(get_parameter_value(myconstants.REPORTS_SECTION_NAME), myconstants.REPORT_FILE_PREFFIX, ".xlsx")
    
    def get_reports_list(self):
        return(self.reports_list)
    
    def get_report_file_name_by_num(self, num):
        return(self.reports_list[num])


def get_parameter_value(paramname):
    # Читаем настройки
    settings_df = pd.read_excel("Settings.xlsx", engine='openpyxl')
    settings_df.dropna(how='all', inplace=True)
    ret_value = settings_df[settings_df["ParameterName"]==paramname]["ParameterValue"].to_list()[0]
    if type(ret_value) == str:
        ret_value = ret_value.replace("\\", "/")
        if ret_value[-1] == "/":
            ret_value = ret_value[:-1:]
    
    return (ret_value)

def load_raw_data(raw_file, ui_handle):
    # Загружаем сырые данные
    ui_handle.set_status("Начинаем загружать исходные данные.")
    df = pd.read_excel(raw_file, engine='openpyxl')
    ui_handle.set_status("Удаляем 'na'.")
    df.dropna(how='all', inplace=True)
    ui_handle.set_status("Переименовываем столбцы и удаляем лишние.")
    df.rename(columns=myconstants.RAW_DATA_COLUMNS, inplace=True)
    exist_drop_columns_list = list(set(myconstants.RAW_DATA_DROP_COLUMNS) & set(df.dtypes.keys()))
    df.drop(columns=exist_drop_columns_list, inplace=True)
    ui_handle.set_status("Исходные данные загружены в слабообработанном виде.")

    return (df)

def load_parameter_table(tablename):
    # Загружаем соответствующу таблицу с параметрами
    parameter_df = pd.read_excel(get_parameter_value(myconstants.PARAMETERS_SECTION_NAME) + "/" + tablename, engine='openpyxl')
    parameter_df.dropna(how='all', inplace=True)
    
    return(parameter_df)

def udata_2_date(data):

    if (data != data):
        ret_date = data
    elif type(data) == str:
        ret_date = dt.datetime.strptime("01." + data, '%d.%m.%Y')
    elif type(data) == float:
        ret_date = dt.datetime.strptime("01." + str(data), '%d.%m.%Y')
    else:
        ret_date = data
    
    return(ret_date)

def calc_fact_fte(FactHour, Northern, CHour, NHour, Project, PlanFTE):
    if Project.find(myconstants.FACT_IS_PLAN_MARKER) >= 0:
        fact_fte = PlanFTE
    else:
        month_hours = NHour if Northern else CHour
        fact_fte = round(FactHour / month_hours, myconstants.ROUND_FTE_VALUE)
    return(fact_fte)

def add_combine_columns(df):
    # "Month",
    # "FN",
    # "Division",
    # "User",
    # "Project",
    # "ProjectType",
    # "ProjectSubType",
    # "PlanFTE",
    # "FactFTE"
    # Подразделение + Проект + ФИО + Месяц
    # Подразделение + ФИО + Проект + Месяц
    # Подразделение + ФИО + Месяц
    # Подразделение + Проект + Месяц
    # ПМ + Проект + Месяц
    # ПМ + Проект + ФИО + Месяц
    df["ShortProject"] = df["Project"].str[:7]
    
    df["FN_Proj"] = df["FN"] + "#" + df["ShortProject"]
    df["FN_Proj_Month"] = df["FN"] + "#" + df["ShortProject"] + "#" + df["Month"]
    
    df["FN_Proj_User"] = df["FN"] + "#" + df["ShortProject"] + "#" + df["User"]
    df["FN_Proj_User_Month"] = df["FN"] + "#" + df["ShortProject"] + "#" + df["User"] + "#" + df["Month"]
    
    df["Pdr_User"] = df["Division"] + "#" + df["User"]
    df["Pdr_User_Month"] = df["Division"] + "#" + df["User"] + "#" + df["Month"]
    
    df["Pdr_User_Proj"] = df["Division"] + "#" + df["User"] + "#" + df["ShortProject"]
    df["Pdr_User_Proj_Month"] = df["Division"] + "#" + df["User"] + "#" + df["ShortProject"] + "#" + df["Month"]
    
    df["ProjMang_Proj"] = df["ProjectManager"] + "#" + df["ShortProject"]
    df["ProjMang_Proj_Month"] = df["ProjectManager"] + "#" + df["ShortProject"] + "#" + df["Month"]
    
    df["ProjMang_Proj_User"] = df["ProjectManager"] + "#" + df["ShortProject"] + "#" + df["User"]
    df["ProjMang_Proj_User_Month"] = df["ProjectManager"] + "#" + df["ShortProject"] + "#" + df["User"] + "#" + df["Month"]

def prepare_data(raw_file_name, p_delete_vacation, ui_handle):
    data_df = load_raw_data(raw_file_name, ui_handle)
    
    month_hours_df = load_parameter_table(myconstants.MONTH_WORKING_HOURS_TABLE)
    divisions_names_df = load_parameter_table(myconstants.DIVISIONS_NAMES_TABLE)
    fns_names_df = load_parameter_table(myconstants.FNS_NAMES_TABLE)
    projects_sub_types_df = load_parameter_table(myconstants.PROJECTS_SUB_TYPES_TABLE)
    ui_handle.set_status(f"Загружены таблицы с параметрами (всего строк данных: {data_df.shape[0]})")

    for column_name in set(data_df.dtypes.keys()) - set(myconstants.DONT_REPLACE_ENTER):
        if data_df.dtypes[column_name] == type(str):
            data_df[column_name] = data_df[column_name].str.replace("\n", "")
    ui_handle.set_status(f"Удалены переносы строк (всего строк данных: {data_df.shape[0]})")
    
    data_df["FDate"] = data_df["FDate"].apply(lambda param: udata_2_date(param))
    ui_handle.set_status(f"Обновлён формат данных даты первого дня месяца (всего строк данных: {data_df.shape[0]})")

    data_df['Northern'].replace(myconstants.BOOLEAN_VALUES_SUBST, inplace=True)
    data_df = data_df.merge(month_hours_df, left_on="FDate", right_on="FirstDate", how="inner")
    ui_handle.set_status(f"Проведено объединение с таблицей с рабочими часами (всего строк данных: {data_df.shape[0]})")
    
    ui_handle.set_status("... начинаем пересчет фактичеких часов в FTE.")
    data_df["FactFTE"] = \
        data_df[["FactHour", "Northern", "CHour", "NHour", "Project", "PlanFTE"]].apply( \
            lambda param: calc_fact_fte(*param), axis=1)
    ui_handle.set_status(f"Пересчитано (всего строк данных: {data_df.shape[0]})")

    data_df = data_df.merge(divisions_names_df, left_on="DivisionRaw", right_on="FullDivisionName", how="left")
    ui_handle.set_status(f"Выполнено объединение с таблицей с подразделениями (всего строк данных: {data_df.shape[0]})")
    ui_handle.set_status("... ищем пустые и восстанавливаем.")
    data_df["Division"] = data_df[["ShortDivisionName", "DivisionRaw"]].apply(lambda param: param[1] if pd.isna(param[0]) else param[0], axis=1)
    ui_handle.set_status(f"Все подразделенния заполнены (всего строк данных: {data_df.shape[0]})")

    data_df = data_df.merge(fns_names_df, left_on="FNRaw", right_on="FullFNName", how="left")
    data_df["FN"] = data_df[["ShortFNName", "FNRaw"]].apply(lambda param: param[1] if pd.isna(param[0]) else param[0], axis=1)
    ui_handle.set_status(f"Данные объединены с таблицей с ФН (всего строк данных: {data_df.shape[0]})")
    
    data_df["ProjectType"] = \
        data_df[["Project", "ProjectType"]].apply(
            lambda param: "S" if param[0].find(myconstants.FACT_IS_PLAN_MARKER) >= 0 else param[1], axis=1)
    ui_handle.set_status(f"Уточнены типы проектов (всего строк данных: {data_df.shape[0]})")

    data_df = data_df.merge(projects_sub_types_df, left_on="Project", right_on="ProjectName", how="left")
    data_df["ProjectSubType"] = \
        data_df[["ProjectType", "ProjectSubTypePart"]].apply(
            lambda param: param[0] + myconstants.OTHER_PROJECT_SUB_TYPE if pd.isna(param[1]) else param[1], axis=1)
    ui_handle.set_status(f"... и типы ПОДпроектов (всего строк данных: {data_df.shape[0]})")

    if p_delete_vacation:
        vacancy_text = myconstants.VACANCY_NAME_TEXT
        vacancy_text = vacancy_text.lower()
        data_df["User"] = \
            data_df["User"].apply(
                lambda param: vacancy_text if param.replace(" ", "").lower()[:len(vacancy_text)]==vacancy_text else param)
        
        data_df = data_df[data_df["User"] != vacancy_text]
        ui_handle.set_status(f"Удалены вакансии (всего строк данных: {data_df.shape[0]})")
    
   
    add_combine_columns(data_df)
    ui_handle.set_status(f"Добавленны производные столбцы (конкатинация) (всего строк данных: {data_df.shape[0]})")
    
    return(data_df[myconstants.RESULT_DATA_COLUMNS])

@thread
def send_df_2_xls(report_file_name, raw_file_name, ui_handle):
    p_delete_vacation = load_param(myconstants.PARAMETER_SAVED_VALUE_DELETE_VAC, True)
    p_open_in_excel = load_param(myconstants.PARAMETER_SAVED_VALUE_OPEN_IN_EXCEL, False)
    
    save_param(myconstants.PARAMETER_SAVED_SELECTED_REPORT,ui_handle.reports_list.index(report_file_name) + 1)
    
    myconstants.ROUND_FTE_VALUE = get_parameter_value(myconstants.ROUND_FTE_SECTION_NAME)
    
    report_prepared_name = \
        get_parameter_value(myconstants.REPORTS_PREPARED_SECTION_NAME) + "/" + \
        report_file_name + raw_file_name + myconstants.EXCEL_FILES_ENDS

    report_file_name = \
        get_parameter_value(myconstants.REPORTS_SECTION_NAME) + "/" + \
        myconstants.REPORT_FILE_PREFFIX + \
        report_file_name + \
        myconstants.EXCEL_FILES_ENDS
    
    raw_file_name = \
        get_parameter_value(myconstants.RAW_DATA_SECTION_NAME) + "/" + \
        raw_file_name + \
        myconstants.EXCEL_FILES_ENDS
    
    ui_handle.clear_status()
    ui_handle.set_status(myconstants.TEXT_LINES_SEPARATOR)
    ui_handle.set_status(f"Выбран отчет: {report_file_name}")
    ui_handle.set_status(f"файл с данными: {raw_file_name}")
    ui_handle.set_status(f"Вакансии : {'удалить из отчета.' if p_delete_vacation else 'оставить в отчете.'}")
    ui_handle.set_status(myconstants.TEXT_LINES_SEPARATOR)

    ui_handle.set_status("Проверим структуру файла с формой отчёта.")
    wb = load_workbook(report_file_name)
    if (myconstants.RAW_DATA_SHEET_NAME not in wb.sheetnames):
        ui_handle.set_status("")
        ui_handle.set_status("")
        ui_handle.set_status("[Ошибка в структуре отчета]")
        ui_handle.set_status("")
        ui_handle.set_status("В файле для выбранной формы отчёта отсутствует необходимый лист для данных.")
        ui_handle.set_status("Формирование отчёта не возможно.")
        save_param(myconstants.PARAMETER_FILENAME_OF_LAST_REPORT, "")
        ui_handle.enable_buttons()
        return
    elif (myconstants.UNIQE_LISTS_SHEET_NAME not in wb.sheetnames):
        ui_handle.set_status("")
        ui_handle.set_status("")
        ui_handle.set_status("[Ошибка в структуре отчета]")
        ui_handle.set_status("")
        ui_handle.set_status("В файле для выбранной формы отчёта отсутствует необходимый лист для уникальных списков.")
        ui_handle.set_status("Формирование отчёта не возможно.")
        save_param(myconstants.PARAMETER_FILENAME_OF_LAST_REPORT, "")
        ui_handle.enable_buttons()
        return
    else:
        ui_handle.set_status("Ошибок не найдено.")

    ui_handle.set_status("Файл Excel с формой отчёта подгружен.")
    
    report_df = prepare_data(raw_file_name, p_delete_vacation, ui_handle)
    ui_handle.set_status(f"Таблица для загрузки полностью подготовлена (всего строк данных: {report_df.shape[0]})")

    ui_handle.set_status("Начинаем перенос строк в Excel:")
    ui_handle.set_status("--")
    counter = 0
    str_rows_in_df = " --> " + str(report_df.shape[0])
    for one_row in dataframe_to_rows(report_df, index=False, header=False):
        counter += 1
        ui_handle.change_last_status_line(str(counter) + str_rows_in_df)
        wb[myconstants.RAW_DATA_SHEET_NAME].append(one_row)

    ui_handle.set_status("Строки в Excel скопированы.")

    ui_handle.set_status("Формируем списки с уникальными значениями.")
    # Запоним списки уникальными значениями
    column = 0
    values_dict = dict()
    while True:
        try:
            element = wb[myconstants.UNIQE_LISTS_SHEET_NAME][1][column].value
            values_dict[element] = column
            column += 1
        except:
            break

    full_column_list = report_df.columns.tolist()
    columns_4_unique_list = [column for column in values_dict.keys() if column in full_column_list]
    if len(columns_4_unique_list)==0:
        ui_handle.set_status("")
        ui_handle.set_status("")
        ui_handle.set_status("[Ошибка в структуре отчета]")
        ui_handle.set_status("")
        ui_handle.set_status("В файле для выбранной формы на листе для уникальных списков не указано ничего.")
        ui_handle.set_status("Формирование отчёта остановлено.")
        save_param(myconstants.PARAMETER_FILENAME_OF_LAST_REPORT, "")
        ui_handle.enable_buttons()
        return

    
    ui_handle.set_status(f"Всего списков: {len(columns_4_unique_list)} шт.")
    ui_handle.set_status(f"")
    for index, one_column in enumerate(columns_4_unique_list):
        unique_elements_list = sorted(report_df[one_column].unique())
        ui_handle.change_last_status_line(f"{index+1} из {len(columns_4_unique_list)} ({one_column}): {len(unique_elements_list)}")
        xls_column_num = values_dict[one_column]
        for xls_row, uvalue in enumerate(unique_elements_list):
            wb[myconstants.UNIQE_LISTS_SHEET_NAME][2 + xls_row][xls_column_num].value = uvalue
    ui_handle.change_last_status_line("Собраны и сохранениы списки с уникальными значениями.")
                
    ui_handle.set_status(myconstants.TEXT_LINES_SEPARATOR)
    ui_handle.set_status(f"Для сохранения выбран файл: {report_prepared_name}")
    ui_handle.set_status(f"Начинаем сохранение ....")
    try:
        wb.save(report_prepared_name)
        ui_handle.change_last_status_line(f"Данные сохранены.")
        save_param(myconstants.PARAMETER_FILENAME_OF_LAST_REPORT, report_prepared_name)
        if p_open_in_excel:
            ui_handle.set_status(f"Открываем сформированный файл в Excel.")
    except:
        p_open_in_excel = False
        save_param(myconstants.PARAMETER_FILENAME_OF_LAST_REPORT, "")
        ui_handle.set_status("Файл не доступен, возможно, он уже открыт.")
        e = sys.exc_info()[1]

    if p_open_in_excel:
        subprocess.Popen(report_prepared_name, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, stdin=subprocess.PIPE)
    ui_handle.set_status(myconstants.TEXT_LINES_SEPARATOR)
    ui_handle.enable_buttons()


if __name__ == "__main__":
    print(get_files_list(get_parameter_value(myconstants.REPORTS_SECTION_NAME)))